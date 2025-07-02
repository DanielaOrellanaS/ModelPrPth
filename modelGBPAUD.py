import pandas as pd
import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import Dataset, DataLoader
import os
import numpy as np
import pickle
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ====================== Normalización ======================

def normalize(column, min_val, max_val):
    return (column - min_val) / (max_val - min_val)

# ====================== Carga y preprocesamiento ======================

file_path = r'C:\Users\user\OneDrive\Documentos\Trading\ModelPrPth\ModelAndTest\DataFiles\Data_Entrenamiento_GBPAUD.xlsx'
data = pd.read_excel(file_path)
data.columns = data.columns.str.strip()

# Limpieza y conversión de fecha
data['fecha'] = data['fecha'].str.replace(',', '-', regex=False)  # Reemplaza ',' por '-' para formato YYYY-MM-DD
data['fecha'] = pd.to_datetime(data['fecha'], format="%Y-%m-%d %H:%M")

# Extraer componentes temporales
data['dia_semana'] = data['fecha'].dt.weekday        # Lunes=0, Domingo=6
data['hora'] = data['fecha'].dt.hour                 # 0-23
data['minuto'] = data['fecha'].dt.minute             # 0-55, en saltos de 5

# Normalizar variables temporales
data['dia_semana'] = data['dia_semana'] / 6.0
data['hora'] = data['hora'] / 23.0
data['minuto'] = data['minuto'] / 55.0

# Normalización precios 5min
min_precio5, max_precio5 = data['preciolow5'].min(), data['preciohigh5'].max()
min_volume5, max_volume5 = data['volume5'].min(), data['volume5'].max()
data['volume5'] = normalize(data['volume5'], min_volume5, max_volume5)
min_volume15, max_volume15 = data['volume15'].min(), data['volume15'].max()
data['volume15'] = normalize(data['volume15'], min_volume15, max_volume15)

print("MIN VOLUMEN", min_volume5)
print("MAX VOLUMEN", max_volume5)

for col in ['precioopen5', 'precioclose5', 'preciohigh5', 'preciolow5']:
    data[col] = normalize(data[col], min_precio5, max_precio5)

# Normalización precios 15min
min_precio15, max_precio15 = data['preciolow15'].min(), data['preciohigh15'].max()
for col in ['precioopen15', 'precioclose15', 'preciohigh15', 'preciolow15']:
    data[col] = normalize(data[col], min_precio15, max_precio15)

# Normalización RSI y Stochastic (escala 0-1)
for col in ['rsi5', 'rsi15', 'iStochaMain5', 'iStochaSign5', 'iStochaMain15', 'iStochaSign15']:
    data[col] = data[col] / 100.0

# Guardar profit real antes de normalizar
data['profit_original'] = data['profit'].fillna(0)

# Normalización de profit (target)
min_profit = data['profit_original'].min()
max_profit = data['profit_original'].max()
print(f"MIN PROFIT ORIGINAL: {min_profit:.6f}")
print(f"MAX PROFIT ORIGINAL: {max_profit:.6f}")

print("DATOS PROFIT ORIGINAL: ", data['profit_original'])
print("DATOS NORMALIZADOS ANTES DE AGREGAR: ", normalize(data['profit_original'], min_profit, max_profit))

data['profit'] = normalize(data['profit_original'], min_profit, max_profit)

# ====================== Columnas de entrada ======================

input_columns = [
    'dia_semana', 'hora', 'minuto',
    "precioopen5", "precioclose5", "precioclose5",  # duplicado para dar más peso
    "preciohigh5", "preciolow5", "volume5",
    "precioopen15", "precioclose15", "preciohigh15", "preciolow15", "volume15",
    "rsi5", "rsi15", "iStochaMain5", "iStochaSign5", "iStochaMain15", "iStochaSign15"
]

# ====================== Dataset ======================

class TradingDataset(Dataset):
    def __init__(self, df):
        self.X = df[input_columns].values
        self.y = df['profit'].values.reshape(-1, 1) 

    def __len__(self):
        return len(self.X)

    def __getitem__(self, idx):
        return (
            torch.tensor(self.X[idx], dtype=torch.float32),
            torch.tensor(self.y[idx], dtype=torch.float32)
        )

# ====================== Modelo ======================

class TradingModel(nn.Module):
    def __init__(self):
        super().__init__()
        self.net = nn.Sequential(
            nn.Linear(len(input_columns), 64),
            nn.ReLU(),
            nn.Linear(64, 32),
            nn.ReLU(),
            nn.Linear(32, 1)
        )

    def forward(self, x):
        return self.net(x)

# ====================== Entrenamiento ======================

dataset = TradingDataset(data)
dataloader = DataLoader(dataset, batch_size=64, shuffle=True)

model = TradingModel()
criterion = nn.SmoothL1Loss()
optimizer = optim.Adam(model.parameters(), lr=0.001)

epochs = 50
for epoch in range(epochs):
    total_loss = 0.0
    for X_batch, y_batch in dataloader:
        optimizer.zero_grad()
        output = model(X_batch)
        loss = criterion(output, y_batch)
        loss.backward()
        optimizer.step()
        total_loss += loss.item()
    print(f"Epoch {epoch+1}/{epochs} - Loss: {total_loss/len(dataloader):.6f}")

# ====================== Guardado del modelo ======================

os.makedirs("Trading_Model", exist_ok=True)
torch.save(model.state_dict(), "Trading_Model/trading_model_GBPAUD.pth")

# Guardar min/max para normalización/desnormalización posterior

print("MINIMOS Y MAXIMOS VOLUMEN: ", min_volume5, max_volume5)
min_max_dict = {
    "min_profit": min_profit,
    "max_profit": max_profit,
    "min_precio5": min_precio5,
    "max_precio5": max_precio5,
    "min_precio15": min_precio15,
    "max_precio15": max_precio15,
    "min_volume5": min_volume5,
    "max_volume5": max_volume5,
    "min_volume15": min_volume15,
    "max_volume15": max_volume15,
}

with open("Trading_Model/min_max_GBPAUD.pkl", "wb") as f:
    pickle.dump(min_max_dict, f)

# Guardar Excel con todos los datos normalizados
output_path = r"C:\Users\user\OneDrive\Documentos\Trading\ModelPrPth\ModelAndTest\DataFiles\Datos_Entrenamiento_GBPAUD_Normalizado.xlsx"
data.to_excel(output_path, index=False)

# === Pintar celdas normalizadas ===
wb = load_workbook(output_path)
ws = wb.active

# Color rosa personalizado (#D42CA8)
fill = PatternFill(start_color="D42CA8", end_color="D42CA8", fill_type="solid")

# Pintar columnas que fueron normalizadas
cols_normalizadas = [
    'dia_semana', 'hora', 'minuto',
    'precioopen5', 'precioclose5', 'preciohigh5', 'preciolow5',
    'volume5',
    'precioopen15', 'precioclose15', 'preciohigh15', 'preciolow15',
    'volume15',
    'rsi5', 'rsi15',
    'iStochaMain5', 'iStochaSign5', 'iStochaMain15', 'iStochaSign15',
    'profit'
]

for col_name in cols_normalizadas:
    if col_name in data.columns:
        col_idx = list(data.columns).index(col_name) + 1  # Excel usa 1-based indexing
        for row in range(2, ws.max_row + 1):  # desde fila 2 (sin cabecera)
            ws.cell(row=row, column=col_idx).fill = fill

wb.save(output_path)
print(f"Archivo con datos normalizados guardado en: {output_path}")